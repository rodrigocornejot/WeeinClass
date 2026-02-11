import json
import openpyxl
import traceback
import calendar
from django.contrib.auth.models import User, Group
from django.core.paginator import Paginator
from rest_framework.generics import ListAPIView
from rest_framework import viewsets
from django.shortcuts import render, redirect, get_object_or_404
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Alumno, Curso, Nota, Matricula, Clase, Tarea, Area, Evento, UnidadCurso, AsistenciaUnidad, Pago, Cuota, Egresado, Reprogramacion
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django import forms
from .forms import CursoForm, MatriculaAdminForm, MatriculaForm, NotaForm, Pago, ServicioForm
from django.db.models import Count, Avg, Q, Sum
from .forms import AlumnoForm, UsuarioCreateForm, UsuarioUpdateForm
from datetime import date, timedelta, datetime
from django.contrib.auth.forms import AuthenticationForm
from openpyxl.utils import get_column_letter
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.utils import timezone, dateparse
from .utils import obtener_datos_dashboard
from .serializers import MatriculaSerializer
from rest_framework.viewsets import ReadOnlyModelViewSet, ModelViewSet
from django.views import View
from django.contrib.auth.decorators import login_required, user_passes_test
from .utils import crear_asistencias_para_matricula  # Importar la función
import ast
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from django.views.decorators.http import require_POST
from django.utils.dateparse import parse_date
from django.db import transaction, IntegrityError
from cursos.models import Certificado
from django.core.files.base import ContentFile
from django.conf import settings
from django.utils.timezone import now
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from cursos.utils import emitir_certificado_si_corresponde, verificar_y_generar_certificado, puede_generar_certificado
from django.template.loader import render_to_string
import os
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import ServicioExtra, CategoriaServicio, EntregaKit
from django.core.mail import EmailMessage
from math import ceil 
from .decorators import solo_admin, solo_asesora, asesora_o_profesor, en_grupo_403
from .forms_users import CrearUsuarioForm

CURSO_COLORES = {
    "INST": "#00B8D9",   # celeste
    "VDF":  "#2ECC71",   # verde
    "PLC1": "#FF7A00",   # naranja fuerte
    "PLC2": "#FFD400",   # amarillo
    "REDES":"#8E44AD",   # morado
    "LOGO": "#1F6FEB",   # azul
    "ELEC": "#E74C3C",   # rojo
}


CODIGO_POR_CURSO = {
    'INSTRUMENTACION INDUSTRIAL': 'INST',
    'PLC NIVEL 1': 'PLC1',
    'PLC NIVEL 2': 'PLC2',
    'REDES INDUSTRIALES': 'REDES',
    'VARIADORES DE FRECUENCIA': 'VDF',
    'PLC LOGO! V8': 'PLC_LOGO',
    'AUTOMATIZACION PARA EL ARRANQUE DE MOTORES ELECTRICOS CON PLC LOGO V8!': 'ELEC_LOGO',
    'AUTOMATIZACION INDUSTRIAL CON PLC LOGO V8!': 'LOGO',
}

DIAS_SEMANA = {
    0: 'lunes',
    1: 'martes',
    2: 'miercoles',
    3: 'jueves',
    4: 'viernes',
    5: 'sabado',
    6: 'domingo'
}

MODALIDAD_CHOICES = (
    ('extendida', 'Extendida'),
    ('full_day', 'Full Day'),
)

MODALIDAD_CONFIG = {
    'full_day': {
        'num_weeks': 3,
        'days_of_week': ['domingo'],
        'hora_inicio': '09:00',
        'hora_fin': '18:00',
    },
    'intensivo': {
        'num_weeks': 3,
        'days_of_week': ['lunes','miercoles','viernes'],
        'hora_inicio': '09:00',
        'hora_fin': '13:00',
    },
    'extendida': {
        'num_weeks': 6,
        'days_of_week': ['sabado'],
        'hora_inicio': '09:00',
        'hora_fin': '13:00',
    },
}

SESIONES_POR_CURSO = {
    "INST": 6,
    "PLC1": 6,
    "PLC2": 6,
    "REDES": 5,
    "VDF": 6,
}

CURSO_CONFIG = {
    "INST":      {"codigo": "INST",      "sesiones": 6, "color": "#06b6d4"},
    "PLC1":      {"codigo": "PLC1",      "sesiones": 6, "color": "#f97316"},
    "PLC2":      {"codigo": "PLC2",      "sesiones": 6, "color": "#eab308"},
    "REDES":     {"codigo": "REDES",     "sesiones": 5, "color": "#a855f7"},
    "VDF":       {"codigo": "VDF",       "sesiones": 6, "color": "#22c55e"},
    "ELEC_LOGO": {"codigo": "ELEC_LOGO", "sesiones": 5, "color": "#ef4444"},
    "LOGO":      {"codigo": "LOGO",      "sesiones": 5, "color": "#2563eb"},
}

import unicodedata

def exportar_excel(titulo: str, columnas: list[str], filas: list[list], filename: str = "reporte.xlsx"):
    """
    columnas: ["Col1", "Col2", ...]
    filas: [[v1, v2, ...], [v1, v2, ...]]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append([titulo])
    ws.append([])

    ws.append(columnas)

    for row in filas:
        ws.append(row)

    # Ajustar ancho columnas
    for i, col in enumerate(columnas, start=1):
        max_len = len(str(col))
        for r in range(4, 4 + len(filas)):
            val = ws.cell(row=r, column=i).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def _norm(txt: str) -> str:
    txt = (txt or "").strip().upper()
    # quitar tildes
    txt = "".join(
        c for c in unicodedata.normalize("NFD", txt)
        if unicodedata.category(c) != "Mn"
    )
    # limpiar dobles espacios
    txt = " ".join(txt.split())
    return txt

def curso_codigo_y_sesiones(curso_nombre: str, curso_obj=None):
    n = _norm(curso_nombre)

    # ✅ si ya te guardaron el "nombre" como código (VDF, INST, PLC1...)
    if n in ("INST", "VDF", "PLC1", "PLC2", "REDES", "LOGO", "ELEC", "ELEC_LOGO", "PLC_LOGO"):
        # sesiones
        sesiones = CURSO_SESIONES.get(n, 6)
        # PLC_LOGO lo tratamos como LOGO
        if n == "PLC_LOGO":
            return "LOGO", CURSO_SESIONES.get("LOGO", 5)
        # ELEC_LOGO lo tratamos como ELEC (color rojo)
        if n == "ELEC_LOGO":
            return "ELEC", 5
        return n if n != "ELEC_LOGO" else "ELEC", sesiones

    # ✅ por nombre largo
    if "INSTRUMENTACION" in n:
        return "INST", 6

    if "VARIADORES" in n or "VDF" in n:
        return "VDF", 6

    if "PLC" in n and "NIVEL 1" in n:
        return "PLC1", 6

    if "PLC" in n and "NIVEL 2" in n:
        return "PLC2", 6

    if "REDES" in n:
        return "REDES", 5
    if "ARRANQUE" in n and "LOGO" in n:
        return "ELEC", 5

    if "LOGO" in n:
        # incluye PLC LOGO
        return "LOGO", 5

    print("⚠️ Curso no reconocido:", curso_nombre)
    return "INST", 6

@login_required
@asesora_o_profesor
def calendario_matriculas(request):
    start_str = (request.GET.get("start") or "")[:10]
    end_str   = (request.GET.get("end") or "")[:10]

    clases = (
        Clase.objects
        .select_related("curso")
        .prefetch_related("matriculas__alumno")
        .order_by("fecha")
    )

    try:
        if start_str and end_str:
            start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
            end_date   = datetime.strptime(end_str, "%Y-%m-%d").date()
            clases = clases.filter(fecha__gte=start_date, fecha__lt=end_date)
    except ValueError:
        pass

    eventos = []

    for clase in clases:
        nombre_curso = (clase.curso.nombre or "").strip()

        # ✅ usa curso_obj para que respete curso.codigo si existe
        codigo, _ = curso_codigo_y_sesiones(nombre_curso, curso_obj=clase.curso)

        color = CURSO_COLORES.get(codigo, "#0d6efd")

        # si no hay matrículas, no hay evento
        if not clase.matriculas.exists():
            continue

        for matricula in clase.matriculas.all():
            eventos.append({
                "id": f"clase-{clase.id}-mat-{matricula.id}",
                "title": f"{codigo} - {matricula.alumno.nombre}",
                "start": clase.fecha.strftime("%Y-%m-%d"),
                "color": color,
                "textColor": "black",

                # ✅ ESTO ES LO QUE TE FALTABA (para el click)
                "extendedProps": {
                    "clase_id": clase.id,
                    "matricula_id": matricula.id,
                    "curso": clase.curso.nombre,
                    "alumno": matricula.alumno.nombre,
                }
            })

    print("EVENTOS DEVUELTOS:", len(eventos))
    return JsonResponse(eventos, safe=False)



# Formulario de login (puedes usar el AuthenticationForm de Django, pero aquí un ejemplo sencillo)
class LoginForm(forms.Form):
    username = forms.CharField(label="Nombre de usuario")
    password = forms.CharField(widget=forms.PasswordInput, label="Contraseña")

@csrf_protect
def mi_login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            grupos = list(user.groups.values_list("name", flat=True))
            print("Usuario autenticado:", user.username)
            print("Grupos:", grupos)

            # ✅ ADMIN (soporta ambos nombres)
            if user.is_superuser or user.groups.filter(name__in=["Admin", "Administradores"]).exists():
                print("Redirigiendo a: menu_admin (admin)")
                return redirect("menu_admin")

            # ✅ ASESORA
            if user.groups.filter(name="Asesora").exists():
                print("Redirigiendo a: menu_admin (asesora)")
                return redirect("menu_admin")

            # ✅ PROFESOR (soporta ambos nombres)
            if user.groups.filter(name__in=["Profesor", "Profesores"]).exists():
                print("Redirigiendo a: menu_admin (profesor)")
                return redirect("menu_admin")

            # ✅ MARKETING
            if user.groups.filter(name="Marketing").exists():
                print("Redirigiendo a: menu_admin (marketing)")
                return redirect("menu_admin")

            # ✅ fallback (si no tiene grupo)
            print("Redirigiendo a: menu_admin (fallback)")
            return redirect("menu_admin")

        else:
            print("Formulario invalido")
            print(form.errors)
            return render(request, "registration/login.html", {"form": form})

    else:
        form = AuthenticationForm()

    return render(request, "registration/login.html", {"form": form})

def es_profesor_o_recepcion(user):
    # Ajusta los nombres de grupos si los tuyos son distintos
    return (
        user.is_superuser or 
        user.groups.filter(name__in=['Profesores', 'Recepcion']).exists()
    )

def lista_alumnos(request):
    alumnos = Alumno.objects.all()
    return render(request, 'cursos/lista_alumnos.html', {
        'alumnos': alumnos    
    })

def es_profesor(user):
    return user.is_superuser or user.groups.filter(name='Profesores').exists()

@login_required
@asesora_o_profesor
def vista_calendario(request):
    cursos = Curso.objects.all()

    # Crear una lista con los nombres de los cursos y sus colores
    cursos_colores = [(curso.nombre, CURSO_COLORES.get(curso.codigo, '#000000')) for curso in cursos]

    # Pasar la lista de cursos con sus colores al contexto de la plantilla
    context = {
        'cursos_colores': cursos_colores,
    }
    return render(request, 'cursos/calendario.html')

def es_admin(user):
    return user.is_superuser or user.groups.filter(name='Administradores').exists()

@login_required
def menu_admin(request):
    user = request.user

    es_admin = (
        user.is_superuser
        or user.groups.filter(name__in=["Admin", "Administradores"]).exists()
    )

    es_profesor = user.groups.filter(name__in=["Profesor", "Profesores"]).exists()

    es_asesora = user.groups.filter(name__in=["Asesora", "Recepcion", "Recepción"]).exists()

    es_marketing = user.groups.filter(name="Marketing").exists()

    # fallback
    if not (es_admin or es_profesor or es_asesora or es_marketing):
        es_asesora = True

    print("🧪 DEBUG MENU_ADMIN")
    print("user:", user.username)
    print("is_superuser:", user.is_superuser)
    print("groups:", list(user.groups.values_list("name", flat=True)))
    print("flags:", {
        "es_admin": es_admin,
        "es_profesor": es_profesor,
        "es_asesora": es_asesora,
        "es_marketing": es_marketing,
    })

    return render(request, "cursos/menu_admin.html", {
        "es_admin": es_admin,
        "es_profesor": es_profesor,
        "es_asesora": es_asesora,
        "es_marketing": es_marketing,
    })

@login_required
@solo_asesora
def crear_curso(request):
    if request.method == 'POST':
        print("Datos recibidos en POST:", request.POST)

        form = CursoForm(request.POST)
        if form.is_valid():

            nombre = form.cleaned_data['nombre']
            fecha = form.cleaned_data.get('fecha')
            horario = form.cleaned_data.get('horario')
            duracion = form.cleaned_data.get('duracion')
            profesor = form.cleaned_data.get('profesor')

            if not nombre or not fecha or not horario or not duracion or not profesor:
                messages.error(request, "Faltan datos obligatorios para el curso.")
                return render(request, 'cursos/crear_curso.html', {'form': form})

            # ✅ Evitar duplicado por nombre+fecha (tu lógica)
            curso_existente = Curso.objects.filter(nombre=nombre, fecha=fecha).first()
            if curso_existente:
                messages.warning(request, "Ese curso ya existe para esa fecha.")
                return redirect('lista_cursos')

            # ✅ Generar código base por nombre
            codigo_base = CODIGO_POR_CURSO.get(nombre)

            if not codigo_base:
                codigo_base, _ = curso_codigo_y_sesiones(nombre)

            # LOGO/PLC_LOGO y ELEC_LOGO/ELEC
            if codigo_base == "PLC_LOGO":
                codigo_base = "LOGO"
            if codigo_base == "ELEC_LOGO":
                codigo_base = "ELEC"


            # ✅ Asegurar código único: si existe, agregar sufijo -2, -3...
            codigo_final = codigo_base
            i = 2
            while Curso.objects.filter(codigo=codigo_final).exists():
                codigo_final = f"{codigo_base}-{i}"
                i += 1

            try:
                with transaction.atomic():
                    curso = Curso.objects.create(
                        codigo=codigo_final,
                        nombre=nombre,
                        fecha=fecha,
                        horario=horario,
                        duracion=duracion,
                        profesor=profesor
                    )
            except IntegrityError:
                messages.error(request, "❌ No se pudo crear: el código ya existe. Intenta de nuevo.")
                return render(request, 'cursos/crear_curso.html', {'form': form})

            messages.success(request, f"✅ Curso creado: {curso.nombre} ({curso.codigo})")
            return redirect('lista_cursos')

        else:
            print("Errores en el formulario:", form.errors)
            messages.error(request, "Revisa el formulario: hay errores.")
    else:
        form = CursoForm()

    return render(request, 'cursos/crear_curso.html', {'form': form})


@login_required
def lista_cursos(request):
    cursos = Curso.objects.all()
    matriculas = Matricula.objects.all()
    return render(request, 'cursos/lista_cursos.html', {'cursos': cursos})

@login_required
def editar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)

    if request.method == "POST":
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Curso actualizado.")
            return redirect("lista_cursos")
    else:
        form = CursoForm(instance=curso)

    return render(request, "cursos/editar_curso.html", {"form": form, "curso": curso})


@login_required
def eliminar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)

    if request.method == "POST":
        curso.delete()
        messages.success(request, "🗑️ Curso eliminado.")
        return redirect("lista_cursos")

    return render(request, "cursos/eliminar_curso.html", {"curso": curso})

def lista_matriculas(request):
    matriculas_qs = (
        Matricula.objects
        .select_related("alumno", "curso")
        .select_related("kit")  # OneToOne (EntregaKit) -> related_name="kit"
        .order_by("-id")
    )

    cursos = Curso.objects.all().order_by("nombre")

    return render(request, "cursos/lista_matriculas.html", {
        "matriculas": matriculas_qs,  # ✅ sin paginator (DataTables pagina)
        "cursos": cursos
    })

@login_required
@asesora_o_profesor
def registrar_asistencia_unidad(request):
    cursos = Curso.objects.all().order_by("nombre")

    curso_id = request.GET.get("curso") or request.POST.get("curso")
    fecha_str = request.GET.get("fecha") or request.POST.get("fecha")

    curso_seleccionado = None
    fecha_seleccionada = None
    clase = None

    alumnos_unidades = []
    unidades = []

    # ---- Parse fecha ----
    if fecha_str:
        try:
            fecha_seleccionada = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            try:
                fecha_seleccionada = datetime.strptime(fecha_str, "%d/%m/%Y").date()
            except ValueError:
                fecha_seleccionada = None

    if curso_id:
        curso_seleccionado = Curso.objects.filter(id=curso_id).first()

    if curso_seleccionado and fecha_seleccionada:
        # ✅ obligar a que exista clase ese día
        clase = Clase.objects.filter(
            curso=curso_seleccionado,
            fecha=fecha_seleccionada
        ).first()

        if clase:
            matriculas = clase.matriculas.select_related("alumno", "curso").distinct()

            # ✅ unidades reales (5 o 6) + nombre_tema correcto
            unidades = asegurar_unidades_curso(curso_seleccionado)

            # Si no hay unidades configuradas, avisamos
            if not unidades:
                messages.warning(
                    request,
                    f"⚠ El curso '{curso_seleccionado.nombre}' no tiene módulos configurados."
                )

            # ---- POST: guardar ----
            if request.method == "POST" and request.POST.get("guardar_asistencias"):
                with transaction.atomic():
                    for matricula in matriculas:
                        for unidad in unidades:
                            campo = f"asistencia_{matricula.id}_{unidad.id}"
                            presente = request.POST.get(campo) == "on"

                            asistencia = AsistenciaUnidad.objects.filter(
                                matricula=matricula,
                                unidad=unidad
                            ).order_by("fecha").first()

                            if presente:
                                if not asistencia:
                                    AsistenciaUnidad.objects.create(
                                        matricula=matricula,
                                        unidad=unidad,
                                        fecha=fecha_seleccionada,
                                        completado=True
                                    )
                                else:
                                    # ya está completado antes → NO crear duplicado
                                    if not asistencia.completado:
                                        asistencia.completado = True
                                        if not asistencia.fecha:
                                            asistencia.fecha = fecha_seleccionada
                                        asistencia.save(update_fields=["completado", "fecha"])
                            else:
                                # ✅ Opcional recomendado: NO borrar módulos ya completados en días anteriores
                                # Solo borrar si fue marcado hoy (para correcciones del mismo día)
                                if asistencia and asistencia.fecha == fecha_seleccionada:
                                    asistencia.delete()

                messages.success(request, "✅ Asistencias guardadas correctamente.")

                for matricula in matriculas:
                    verificar_y_generar_certificado(matricula)

                return redirect(
                    f"/cursos/registrar_asistencia_unidad/?curso={curso_seleccionado.id}&fecha={fecha_seleccionada}"
                )

            # ---- GET: construir tabla ----
            for matricula in matriculas:
                unidades_info = []
                for unidad in unidades:
                    asistencia_global = AsistenciaUnidad.objects.filter(
                        matricula=matricula,
                        unidad=unidad,
                        completado=True
                    ).order_by("fecha").first()

                    marcado_global = asistencia_global is not None
                    marcado_hoy = marcado_global and asistencia_global.fecha == fecha_seleccionada

                    unidades_info.append({
                        "unidad_id": unidad.id,
                        "numero": unidad.numero,
                        "nombre_tema": unidad.nombre_tema,  # ✅ SIEMPRE nombre_tema
                        "completado": marcado_global,     # ✅ ya está completado (cualquier fecha)
                        "marcado_hoy": marcado_hoy,       # ✅ se completó justo hoy
                        "fecha_completado": asistencia_global.fecha if asistencia_global else None,
                    })

                alumnos_unidades.append({
                    "matricula": matricula,
                    "unidades": unidades_info
                })

    context = {
        "cursos": cursos,
        "curso_seleccionado": curso_seleccionado,
        "fecha_seleccionada": fecha_seleccionada,
        "clase": clase,
        "alumnos_unidades": alumnos_unidades,
        "unidades": unidades,
    }
    return render(request, "cursos/registrar_asistencia_unidad.html", context)

def lista_asistencia(request):
    asistencias = AsistenciaUnidad.objects.select_related(
        'matricula__alumno', 'matricula__curso', 'unidad'
    ).order_by('matricula__curso__nombre', 'matricula__alumno__nombre', 'unidad__numero')
    return render(request, 'cursos/lista_asistencia.html', {'asistencias': asistencias})

def ver_asistencia_unidad(request, curso_id):
    # Lógica para obtener las asistencias por unidad
    # Ejemplo:
    asistencias = AsistenciaUnidad.objects.filter(curso_id=curso_id)
    unidades = UnidadCurso.objects.filter(curso_id=curso_id).order_by('id')

    return render(request, 'cursos/ver_asistencia_unidad.html', {
        'asistencias': asistencias,
        'unidades': unidades
    })

def guardar_asistencias(request, curso_id):
    if request.method == 'POST':
        unidades = UnidadCurso.objects.filter(curso_id=curso_id)
        matriculas = Matricula.objects.filter(curso_id=curso_id)

        for matricula in matriculas:
            for unidad in unidades:
                checkbox_name = f'asistencia_{matricula.id}_{unidad.id}'
                completado = checkbox_name in request.POST

                asistencia, created = AsistenciaUnidad.objects.get_or_create(
                    matricula=matricula,
                    unidad=unidad,
                    defaults={'completado': completado}
                )

                if not created:
                    asistencia.completado = completado
                    asistencia.save()

        return redirect('registrar_asistencia_unidad', curso_id=curso_id)

def registrar_nota(request):
    alumnos = Alumno.objects.all()
    context = {'alumnos': alumnos}
    if request.method == 'POST':
        form = NotaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_notas')
    else:
        alumnos = Alumno.objects.filter(matricula__asistencia=True).distinct()
        form = NotaForm()

    return render(request, 'cursos/registrar_nota.html', {'form': form, 'alumnos': alumnos})

def lista_notas(request):
    alumno_id = request.GET.get("alumno_id")  # O request.POST.get("alumno_id") según el caso

    if alumno_id:
        try:
            alumno_id = int(alumno_id)
            notas = Nota.objects.filter(matricula__alumno=alumno_id)
        except ValueError:
            notas = Nota.objects.none()  # Si no es un número válido, devuelve lista vacía
    else:
        notas = Nota.objects.all()  # Opcional: Mostrar todas las notas si no hay filtro

    return render(request, 'cursos/lista_notas.html', {'notas': notas})

def detalle_matricula(request, matricula_id, fecha):
    try:
        matricula = Matricula.objects.select_related('alumno', 'curso').get(id=matricula_id)
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()

        codigo, _ = curso_codigo_y_sesiones(matricula.curso.nombre)

        # ✅ Buscar la clase de ese curso en esa fecha
        clase = (
            Clase.objects
            .filter(curso=matricula.curso, fecha=fecha_dt, matriculas=matricula)
            .first()
        )

        asistencia = None
        if clase:
            asistencia = (
                AsistenciaUnidad.objects
                .select_related("unidad")
                .filter(matricula=matricula, clase=clase)
                .first()
            )

        sesion_num = asistencia.unidad.numero if asistencia else None
        sesion_tema = asistencia.unidad.nombre_tema if asistencia else ""

        return JsonResponse({
            "nombre": matricula.alumno.nombre,
            "curso": matricula.curso.nombre,
            "codigo": codigo,
            "turno": matricula.get_tipo_horario_display(),
            "saldo": float(matricula.saldo_pendiente),
            "fecha": fecha_dt.strftime("%d/%m/%Y"),
            "sesion_num": sesion_num,
            "sesion_tema": sesion_tema,
        })

    except Matricula.DoesNotExist:
        return JsonResponse({"error": "Matrícula no encontrada"}, status=404)

    except Exception as e:
        print("❌ Error:", e)
        return JsonResponse({"error": "Error interno"}, status=500)

TEMAS_POR_CODIGO = {
    "VDF": [
        "SINAMICS",
        "DELTA",
        "SCHNEIDER",
        "ABB",
        "DANFOSS",
        "POWERFLEX",
    ],
    "INST": [
        "INTRODUCCION",
        "TEMPERATURA",
        "PRESION",
        "NIVEL",
        "CAUDAL",
        "VALVULAS",
    ],
    "REDES": [
        "DP",
        "RTU",
        "TCP",
        "PN",
        "ETH/IP",
    ],
}

CURSO_SESIONES = {
    "INST": 6,
    "PLC1": 6,
    "PLC2": 6,
    "VDF": 6,

    "REDES": 5,
    "LOGO": 5,
    "ELEC": 5,
}

@login_required
@solo_asesora
def registrar_matricula(request):
    alumno = None
    dni = request.GET.get('dni') or request.POST.get('dni', '')

    if dni:
        alumno = Alumno.objects.filter(dni=dni).first()

    # =========================
    # GET
    # =========================
    if request.method == "GET":
        form = MatriculaForm(initial={
            'dni': dni,
            'alumno_nombre': alumno.nombre if alumno else ''
        })

        return render(request, 'cursos/registrar_matricula.html', {
            'form': form,
            'alumno': alumno,
            'dni': dni
        })

    # =========================
    # POST
    # =========================
    form = MatriculaForm(request.POST)

    if not alumno:
        form.add_error(None, 'No existe un alumno con ese DNI')

    if not form.is_valid() or not alumno:
        return render(request, 'cursos/registrar_matricula.html', {
            'form': form,
            'alumno': alumno,
            'dni': dni
        })

    # =========================
    # GUARDAR MATRÍCULA
    # =========================
    matricula = form.save(commit=False)
    matricula.registrada_por = request.user
    matricula.alumno = alumno

    precio_base = Decimal(matricula.costo_curso or 0)
    descuento_pct = Decimal(matricula.porcentaje or 0)
    anticipo = Decimal(matricula.primer_pago or 0)

    descuento_monto = (precio_base * descuento_pct / Decimal('100')).quantize(Decimal('0.01'))
    precio_final = (precio_base - descuento_monto).quantize(Decimal('0.01'))

    matricula.precio_final = precio_final
    matricula.saldo_pendiente = (precio_final - anticipo).quantize(Decimal('0.01'))

    matricula.save()
    form.save_m2m()
    EntregaKit.objects.get_or_create(matricula=matricula)

    # =========================
    # CUOTAS
    # =========================
    saldo = matricula.saldo_pendiente
    if saldo > 0 and not Cuota.objects.filter(matricula=matricula).exists():
        monto = (saldo / Decimal('2')).quantize(Decimal('0.01'))
        base = matricula.fecha_inicio or datetime.today().date()

        Cuota.objects.create(
            matricula=matricula,
            numero=1,
            monto=monto,
            monto_pagado=Decimal('0.00'),
            fecha_vencimiento=base + timedelta(days=15),
            pagado=False,
        )
        Cuota.objects.create(
            matricula=matricula,
            numero=2,
            monto=monto,
            monto_pagado=Decimal('0.00'),
            fecha_vencimiento=base + timedelta(days=30),
            pagado=False,
        )

        # =========================
        # FECHAS + UNIDADES
        # =========================

        fecha_inicio = form.cleaned_data.get("fecha_inicio")
        dias = form.cleaned_data.get("dias") or []

        raw_tipo = matricula.tipo_horario or ""
        tipo = raw_tipo.lower()

        nombre_curso = (matricula.curso.nombre or "").strip()
        codigo, total_unidades = curso_codigo_y_sesiones(nombre_curso, curso_obj=matricula.curso)


        # FULL DAY
        es_full_day = tipo.startswith("full")

        sesiones_por_clase = 2 if es_full_day else 1
        total_clases = ceil(total_unidades / sesiones_por_clase)


        # Unidades
        unidades = list(
            UnidadCurso.objects.filter(curso=matricula.curso).order_by("numero")
        )

        if not unidades:
            unidades = asegurar_unidades_curso(matricula.curso)

        unidades = [u for u in unidades if u.numero <= total_unidades]


        fechas = []


        # =========================
        # PERSONALIZADO
        # =========================
        if tipo == "personalizado":

            for i in range(1, total_unidades + 1):

                f = request.POST.get(f"sesion_{i}")

                if f:
                    fechas.append(
                        datetime.strptime(f, "%Y-%m-%d").date()
                    )

            fechas = sorted(fechas)[:total_clases]


        # =========================
        # NORMAL
        # =========================
        else:

            if not fecha_inicio:
                form.add_error('fecha_inicio', 'Ingrese fecha de inicio')
                return render(request, 'cursos/registrar_matricula.html', {
                    'form': form,
                    'alumno': alumno,
                    'dni': dni
                })


            if es_full_day:

                fechas = [
                    fecha_inicio + timedelta(days=7 * i)
                    for i in range(total_clases)
                ]


            else:

                if not dias:
                    form.add_error('dias', 'Seleccione días')
                    return render(request, 'cursos/registrar_matricula.html', {
                        'form': form,
                        'alumno': alumno,
                        'dni': dni
                    })


                dias_map = {
                    'lunes': 0, 'martes': 1, 'miercoles': 2,
                    'jueves': 3, 'viernes': 4,
                    'sabado': 5, 'domingo': 6
                }

                dias_num = sorted({dias_map[d] for d in dias})

                cursor = fecha_inicio

                while len(fechas) < total_clases:

                    if cursor.weekday() in dias_num:
                        fechas.append(cursor)

                    cursor += timedelta(days=1)



        # =========================
        # CREAR CLASES
        # =========================

        if not fechas:
            messages.error(request, "No se generaron fechas.")
            return redirect("vista_calendario")


        unidad_idx = 0


        for fecha in fechas:

            clase, _ = Clase.objects.get_or_create(
                curso=matricula.curso,
                fecha=fecha
            )

            clase.matriculas.add(matricula)


            for _ in range(sesiones_por_clase):

                if unidad_idx >= len(unidades):
                    break

                unidad = unidades[unidad_idx]

                AsistenciaUnidad.objects.get_or_create(
                    matricula=matricula,
                    unidad=unidad,
                    clase=clase,
                    defaults={"completado": False}
                )

                unidad_idx += 1


@login_required
def datos_dashboard(request):
    curso_id = (request.GET.get("curso") or "").strip()
    periodo = (request.GET.get("periodo") or "mes").strip()

    data = calcular_dashboard_data(
        curso_id=curso_id or None,
        periodo=periodo
    )

    return JsonResponse({
        "alumnos_por_curso": data["alumnos_por_curso"],
        "total_cobrado": data["total_cobrado"],
        "operaciones": data["operaciones"],
        "ticket_promedio": data["ticket_promedio"],
        "deuda_total": data["deuda_total"],
        "alumnos_con_deuda": data["alumnos_con_deuda"],
        "por_metodo": data.get("por_metodo", []),

        # ✅ nuevo
        "truncados": data["truncados"],
        "tasa_desercion": data["tasa_desercion"],
    })

def export_dashboard_excel(request):
    # Creamos un workbook y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Dashboard"

    # Definimos encabezados
    headers = ["Curso", "Total Alumnos", "Promedio de Notas"]
    ws.append(headers)

    # Consulta de datos: alumnos por curso y promedio de notas
    alumnos_por_curso = Curso.objects.annotate(total_alumnos=Count('alumnos')).values_list('nombre', 'total_alumnos')
    promedio_notas_qs = Nota.objects.values('curso__nombre').annotate(promedio=Avg('calificacion'))
    notas_dict = {n['curso__nombre']: n['promedio'] for n in promedio_notas_qs}

    # Rellenamos la hoja con los datos
    for curso in alumnos_por_curso:
        nombre, total_alumnos = curso
        promedio = notas_dict.get(nombre, 0)
        ws.append([nombre, total_alumnos, promedio])

    # Ajustamos el ancho de las columnas
    for i, column in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Preparamos la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dashboard.xlsx"'
    wb.save(response)
    return response

@login_required
def dashboard(request):
    hoy = timezone.localdate()
    mes_actual = hoy.month
    anio_actual = hoy.year

    # ✅ TOP asesores (mes actual, independiente del filtro)
    top_monto = (
        Pago.objects
        .filter(
            activo=True,
            registrado_por__isnull=False,
            creado_en__year=anio_actual,
            creado_en__month=mes_actual
        )
        .values(
            "registrado_por__id",
            "registrado_por__username",
            "registrado_por__first_name",
            "registrado_por__last_name",
        )
        .annotate(total=Sum("monto"), cantidad=Count("id"))
        .order_by("-total")[:5]
    )

    top_cantidad = (
        Pago.objects
        .filter(
            activo=True,
            registrado_por__isnull=False,
            creado_en__year=anio_actual,
            creado_en__month=mes_actual
        )
        .values(
            "registrado_por__id",
            "registrado_por__username",
            "registrado_por__first_name",
            "registrado_por__last_name",
        )
        .annotate(cantidad=Count("id"), total=Sum("monto"))
        .order_by("-cantidad")[:5]
    )

    cursos = Curso.objects.all()

    # ✅ carga inicial: mes, sin filtro curso
    data = calcular_dashboard_data(periodo="mes")

    context = {
        "top_monto": top_monto,
        "top_cantidad": top_cantidad,
        "cursos": cursos,

        "alumnos_por_curso": data["alumnos_por_curso"],
        "total_cobrado": data["total_cobrado"],
        "operaciones": data["operaciones"],
        "ticket_promedio": data["ticket_promedio"],
        "deuda_total": data["deuda_total"],
        "alumnos_con_deuda": data["alumnos_con_deuda"],
        "periodo": data["periodo"],
    }
    return render(request, "cursos/dashboard.html", context)


@solo_asesora
def registrar_alumnos(request):
    if request.method == 'POST':
        nombre=request.POST.get('nombre')
        dni=request.POST.get('dni')
        telefono=request.POST.get('telefono')
        correo=request.POST.get('correo')

        grado_academico=request.POST.get('grado_academico')
        carrera=request.POST.get('carrera')
        trabajo=request.POST.get('trabajo')
        referencia=request.POST.get('referencia')

        edad_raw = request.POST.get('edad')
        edad = int(edad_raw) if edad_raw else None

        sexo=request.POST.get('sexo')

        distrito=request.POST.get('distrito')
        departamento=request.POST.get('departamento')
        pais=request.POST.get('pais')

        uso_imagen=request.POST.get('uso_imagen') == 'on'
        
        # VALIDACIÓN DE CORREO ÚNICO
        if correo and Alumno.objects.filter(correo=correo).exists():
            return render(request, 'cursos/registrar_alumnos.html', {
                'error': 'El correo ya está registrado'
            })

        if dni and Alumno.objects.filter(dni=dni).exists():
            return render(request, 'cursos/registrar_alumnos.html', {
                'error': 'El DNI ya está registrado'
            })
        
        Alumno.objects.create(
            nombre=nombre,
            dni=dni,
            correo=correo,
            telefono=telefono,
            edad=edad,
            sexo=sexo,
            grado_academico=grado_academico,
            carrera=carrera,
            trabajo=trabajo,
            referencia=referencia,
            pais=pais,
            departamento=departamento,
            distrito=distrito,
            uso_imagen=uso_imagen
        )

        messages.success(request, '✅ Alumno registrado correctamente')
        return redirect('lista_alumnos')

    return render(request, 'cursos/registrar_alumnos.html')

def calcular_dashboard_data(curso_id=None, periodo="mes"):
    fi, ff = rango_periodo(periodo)

    cursos_qs = Curso.objects.all()
    if curso_id:
        cursos_qs = cursos_qs.filter(id=curso_id)

    # =========================
    # A) Matrículas del periodo
    # =========================
    matriculas = Matricula.objects.filter(curso__in=cursos_qs)

    if fi and ff:
        matriculas = matriculas.filter(fecha_inscripcion__range=[fi, ff])

    # ✅ alumnos por curso (cuenta alumnos distintos por curso)
    alumnos_por_curso = (
        matriculas
        .values("curso__nombre")
        .annotate(total_alumnos=Count("alumno_id", distinct=True))
        .order_by("curso__nombre")
    )

    alumnos_por_curso = [
        {"nombre": x["curso__nombre"], "total_alumnos": x["total_alumnos"]}
        for x in alumnos_por_curso
    ]

    # =========================
    # B) Pagos del periodo
    # =========================
    pagos = Pago.objects.filter(activo=True)

    # ✅ filtro por curso si se selecciona (incluye servicios sin matrícula)
    if curso_id:
        pagos = pagos.filter(Q(matricula__curso_id=curso_id) | Q(matricula__isnull=True))

    # ✅ rango de fechas: fecha_pago_real o fallback creado_en
    if fi and ff:
        pagos = pagos.filter(
            Q(fecha_pago_real__range=[fi, ff]) |
            Q(fecha_pago_real__isnull=True, creado_en__date__range=[fi, ff])
        )

    total_cobrado = pagos.aggregate(t=Sum("monto"))["t"] or Decimal("0.00")
    operaciones = pagos.count()
    ticket_promedio = (total_cobrado / operaciones) if operaciones else Decimal("0.00")

    # =========================
    # C) Deuda del periodo
    # =========================
    deuda_total = matriculas.aggregate(t=Sum("saldo_pendiente"))["t"] or Decimal("0.00")
    alumnos_con_deuda = matriculas.filter(saldo_pendiente__gt=0).count()

    # =========================
    # D) Cobros por método
    # =========================
    por_metodo_qs = (
        pagos
        .values("metodo_pago")
        .annotate(total=Sum("monto"), ops=Count("id"))
        .order_by("-total")
    )

    metodo_map = dict(getattr(Pago, "METODO_PAGO_CHOICES", []))
    por_metodo = []
    for x in por_metodo_qs:
        key = x["metodo_pago"] or "—"
        por_metodo.append({
            "metodo": metodo_map.get(key, key),
            "total": float(x["total"] or 0),
            "ops": int(x["ops"] or 0),
        })

    # =========================
    # E) Truncados (nuevo)
    # =========================
    truncados = matriculas.filter(estado="truncado").count()
    activas = matriculas.filter(estado="activa").count()
    finalizadas = matriculas.filter(estado="finalizada").count()

    total_estado = truncados + activas + finalizadas
    tasa_desercion = (truncados * 100 / total_estado) if total_estado else 0

    return {
        "periodo": periodo,
        "fi": fi,
        "ff": ff,

        "alumnos_por_curso": alumnos_por_curso,

        "total_cobrado": f"{total_cobrado:.2f}",
        "operaciones": operaciones,
        "ticket_promedio": f"{ticket_promedio:.2f}",

        "deuda_total": f"{deuda_total:.2f}",
        "alumnos_con_deuda": alumnos_con_deuda,

        "por_metodo": por_metodo,

        # ✅ nuevos
        "truncados": truncados,
        "tasa_desercion": f"{tasa_desercion:.2f}",
    }

def kanban(request):
    ahora = timezone.now()
    user = request.user

    # Obtener área según grupo
    if user.groups.filter(name='Marketing').exists():
        area_name = 'Marketing'
    elif user.groups.filter(name='Profesores').exists():
        area_name = 'Profesores'
    else:
        area_name = None

    # Tareas pendientes y no vencidas
    tareas_pendientes = Tarea.objects.filter(completado=False, fecha_vencimiento__gte=ahora)
    # Tareas vencidas y no completadas
    tareas_vencidas = Tarea.objects.filter(completado=False, fecha_vencimiento__lt=ahora)
    # Tareas completadas
    tareas_completadas = Tarea.objects.filter(completado=True)

    if area_name:
        tareas_pendientes = tareas_pendientes.filter(area_asignada__nombre=area_name)
        tareas_vencidas = tareas_vencidas.filter(area_asignada__nombre=area_name)
        tareas_completadas = tareas_completadas.filter(area_asignada__nombre=area_name)

    return render(request, 'cursos/kanban.html', {
        'tareas_pendientes': tareas_pendientes,
        'tareas_vencidas': tareas_vencidas,
        'tareas_completadas': tareas_completadas,
    })

@login_required
def reprogramar_tarea(request, tarea_id):
    tarea = get_object_or_404(Tarea, id=tarea_id)

    if request.method == 'POST':
        nueva_fecha = request.POST.get('fecha_vencimiento')
        tarea.fecha_vencimiento = timezone.make_aware(
            datetime.strptime(nueva_fecha, '%Y-%m-%d'),
            timezone.get_current_timezone()
        )

        # Si la nueva fecha es futura, quitar la marca de vencida
        if tarea.fecha_vencimiento > timezone.now():
            tarea.vencida = False
        tarea.save()
        return redirect('tareas_vencidas')

    return render(request, 'cursos/reprogramar_tarea.html', {'tarea': tarea})

@login_required
def actualizar_tareas_vencidas(request):
    ahora = timezone.now()
    Tarea.objects.filter(fecha_vencimiento__lt=ahora, vencida=False).update(vencida=True)
    return redirect('kanban')

@login_required
def tareas_vencidas(request):
    if request.user.groups.filter(name='Marketing').exists():
        tareas = Tarea.objects.filter(vencida=True, area_asignada__nombre='Marketing')
    elif request.user.groups.filter(name='Profesores').exists():
        tareas = Tarea.objects.filter(vencida=True, area_asignada__nombre='Profesores')
    else:
        tareas = Tarea.objects.filter(vencida=True)

    return render(request, 'cursos/tareas_vencidas.html', {'tareas': tareas})

def crear_tarea(request):
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        descripcion = request.POST.get('descripcion')
        prioridad = request.POST.get('prioridad')
        tiempo_estimado = request.POST.get('tiempo_estimado')
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        area_asignada = request.POST.get('area_asignada')

        fecha_vencimiento = timezone.make_aware(datetime.strptime(fecha_vencimiento, '%Y-%m-%d'), timezone.get_current_timezone())


        tarea = Tarea.objects.create(
            titulo=titulo,
            descripcion=descripcion,
            prioridad=prioridad,
            tiempo_estimado=tiempo_estimado,
            fecha_vencimiento=fecha_vencimiento,
            area_asignada=Area.objects.get(id=area_asignada)
        )
        return redirect('kanban')
    areas = Area.objects.all()
    return render(request, 'cursos/crear_tarea.html', {'areas': areas})

@csrf_exempt
def actualizar_estado_tarea(request, tarea_id):
    try:
        tarea = Tarea.objects.get(id=tarea_id)
        data = json.loads(request.body)
        tarea_estado = data.get('estado')
        tarea.estado = tarea_estado

        tarea.completado = (tarea_estado == 'Completado')

        tarea.save()
        return JsonResponse({'status': 'success'})
    except Tarea.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'tarea no encontrada'}, status=404)

@login_required
def exportar_tareas_completadas(request):
    tareas = Tarea.objects.filter(estado='completada', vencida=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tareas Completadas"

    headers = ["Título", "Descripción", "Fecha Vencimiento", "Área"]
    ws.append(headers)

    for tarea in tareas:
        ws.append([
            tarea.titulo,
            tarea.descripcion,
            tarea.fecha_vencimiento.strftime("%Y-%m-%d %H:%M"),
            tarea.area_asignada.nombre
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=tareas_completadas.xlsx'
    wb.save(response)
    return response

@login_required
def tareas_completadas(request):
    ahora = timezone.now()
    if request.user.groups.filter(name='Marketing').exists():
        tareas = Tarea.objects.filter(
            completado=True,
            fecha_vencimiento__lt=ahora,
            area_asignada__nombre='Marketing'
        )
    elif request.user.groups.filter(name='Profesores').exists():
        tareas = Tarea.objects.filter(
            completado=True,
            fecha_vencimiento__lt=ahora,
            area_asignada__nombre='Profesores'
        )
    else:
        tareas = Tarea.objects.filter(completado=True, fecha_vencimiento__lt=ahora)

    return render(request, 'cursos/tareas_completadas.html', {'tareas': tareas})

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='Administradores').exists())
def kanban_admin(request):
    # lógica para ver todas las tareas
    return render(request, "kanban/kanban_admin.html")

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Profesores').exists())
def kanban_profesor(request):
    # lógica para ver solo tareas del área de profesores
    return render(request, "kanban/kanban_profesor.html")

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Marketing').exists())
def kanban_marketing(request):
    # lógica para ver solo tareas del área de marketing
    return render(request, "kanban/kanban_marketing.html")

@login_required
def actualizar_fecha_clase(request, clase_id):

    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "msg": "Método inválido"},
            status=405
        )

    # ================== LEER JSON ==================
    try:
        data = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return JsonResponse(
            {"status": "error", "msg": "JSON inválido"},
            status=400
        )

    nueva_fecha_str = (data.get("nueva_fecha") or "").strip()
    solicitante = (data.get("solicitante") or "WEEIN").strip().upper()
    matricula_id = data.get("matricula_id")

    if not nueva_fecha_str or not matricula_id:
        return JsonResponse(
            {"status": "error", "msg": "Faltan datos"},
            status=400
        )

    # ================== NORMALIZAR ==================
    if solicitante not in ("WEEIN", "ALUMNO"):
        solicitante = "WEEIN"

    try:
        nueva_fecha = datetime.strptime(nueva_fecha_str, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse(
            {"status": "error", "msg": "Fecha inválida"},
            status=400
        )

    # ================== OBTENER OBJETOS ==================
    clase = get_object_or_404(Clase, id=clase_id)
    matricula = get_object_or_404(Matricula, id=matricula_id)

    # ================== VALIDAR RELACIÓN ==================
    if not clase.matriculas.filter(id=matricula.id).exists():
        return JsonResponse(
            {"status": "error", "msg": "La matrícula no pertenece a esta clase"},
            status=400
        )

    fecha_anterior = clase.fecha


    # ================== CONTAR REPROGRAMACIONES ==================

    # Solo del alumno, solo en esta matrícula, solo aplicadas
    reprogramaciones_aplicadas = Reprogramacion.objects.filter(
        matricula=matricula,
        solicitante="ALUMNO",
        pagado=True
    ).count()


    # ================== COBRO DESDE 2DA ==================

    if solicitante == "ALUMNO" and reprogramaciones_aplicadas >= 1:

        # Buscar pendiente existente
        reprog_pendiente = Reprogramacion.objects.filter(
            matricula=matricula,
            clase=clase,
            solicitante="ALUMNO",
            pagado=False,
            monto__gt=0
        ).order_by("-creado_en").first()

        if not reprog_pendiente:

            # Crear nuevo cobro
            reprog_pendiente = Reprogramacion.objects.create(
                matricula=matricula,
                clase=clase,
                fecha_anterior=fecha_anterior,
                fecha_nueva=nueva_fecha,
                solicitante="ALUMNO",
                creado_por=request.user,
                monto=Decimal("50.00"),
                pagado=False,
                pagado_en=None
            )

        else:
            # Actualizar si cambió fecha
            reprog_pendiente.fecha_anterior = fecha_anterior
            reprog_pendiente.fecha_nueva = nueva_fecha
            reprog_pendiente.monto = Decimal("50.00")
            reprog_pendiente.creado_por = request.user

            reprog_pendiente.save(update_fields=[
                "fecha_anterior",
                "fecha_nueva",
                "monto",
                "creado_por"
            ])

        # URL a pasarela
        url_pago = f"{reverse('pagar_reprogramacion')}?reprogramacion_id={reprog_pendiente.id}"

        return JsonResponse({
            "status": "payment_required",
            "msg": "Segunda reprogramación: requiere pago S/50",
            "redirect": url_pago
        })


    # ================== APLICAR GRATIS ==================

    clase.fecha = nueva_fecha
    clase.save(update_fields=["fecha"])


    # ================== GUARDAR HISTORIAL GRATIS ==================

    Reprogramacion.objects.create(
        matricula=matricula,
        clase=clase,
        fecha_anterior=fecha_anterior,
        fecha_nueva=nueva_fecha,
        solicitante=solicitante,
        creado_por=request.user,
        monto=Decimal("0.00"),
        pagado=True,
        pagado_en=timezone.now()
    )


    # ================== RESPUESTA ==================

    return JsonResponse({
        "status": "success",
        "msg": "Clase reprogramada correctamente"
    })
    
def eventos_json(request):
    eventos = Evento.objects.select_related("curso", "alumno").all()
    data = []

    for evento in eventos:
        matricula = Matricula.objects.filter(
            alumno=evento.alumno,
            curso=evento.curso
        ).first()

        data.append({
            "id": evento.id,
            "title": f"{evento.curso.nombre} - {evento.alumno.nombre}",
            "start": evento.fecha.isoformat(),
            "extendedProps": {
                "matricula_id": matricula.id if matricula else None,
                "curso": evento.curso.nombre,
                "alumno": evento.alumno.nombre,
            }
        })

    return JsonResponse(data, safe=False)


def editar_alumno(request, alumno_id):
    alumno = get_object_or_404(Alumno, id=alumno_id)
    if request.method == "POST":
        form = AlumnoForm(request.POST, instance=alumno)
        if form.is_valid():
            form.save()
            return redirect('lista_alumnos')
    else:
        form = AlumnoForm(instance=alumno)
    return render(request, "cursos/editar_alumno.html", {"form": form})

def eliminar_alumno(request, alumno_id):
    alumno = get_object_or_404(Alumno, id=alumno_id)
    alumno.delete()
    return redirect('lista_alumnos')

def editar_matricula(request, matricula_id):
    matricula = get_object_or_404(Matricula, id=matricula_id)
    if request.method == "POST":
        form = MatriculaForm(request.POST, instance=matricula)
        if form.is_valid():
            form.save()
            return redirect('lista_matriculas')
    else:
        form = MatriculaForm(instance=matricula)
    return render(request, "cursos/editar_matricula.html", {"form": form})

@login_required
def eliminar_matricula(request, matricula_id):
    matricula = get_object_or_404(Matricula, id=matricula_id)

    # ✅ IMPORTANTÍSIMO: convertir a lista ANTES de borrar
    clases = list(matricula.clase_set.all())

    # Borrar matrícula
    matricula.delete()

    # Borrar clases que quedaron sin matrículas
    for clase in clases:
        if clase.matriculas.count() == 0:
            clase.delete()

    return redirect('lista_matriculas')

@login_required
@solo_asesora
def pagina_pagos(request):
    dni = (request.GET.get("dni") or "").strip()
    matricula_id = (request.GET.get("matricula_id") or "").strip()

    alumno = None
    matriculas = []
    matricula = None
    cuotas = []
    total_pagado_cuotas = Decimal("0.00")

    # ✅ NUEVO: reprogramaciones
    reprogramaciones_pendientes = []
    total_reprogramaciones_pendientes = Decimal("0.00")
    saldo_con_reprogramaciones = None

    if dni:
        alumno = Alumno.objects.filter(dni=dni).first()

        if alumno:
            matriculas = Matricula.objects.filter(alumno=alumno).order_by("-fecha_inscripcion", "-id")

            if matriculas.exists():
                # 1) usar la seleccionada
                if matricula_id:
                    matricula = matriculas.filter(id=matricula_id).first()

                # 2) si no hay seleccionada, tomar la primera
                if not matricula:
                    matricula = matriculas.first()

                # ✅ asegurar cuotas para la matrícula seleccionada
                asegurar_cuotas(matricula)

                # cuotas reales
                cuotas = Cuota.objects.filter(matricula=matricula).order_by("numero", "id")

                # total pagado en cuotas
                total_pagado_cuotas = cuotas.aggregate(total=Sum("monto_pagado"))["total"] or Decimal("0.00")

                # recalcular saldo pendiente coherente (solo CURSO)
                precio_final = getattr(matricula, "precio_final", None)
                if precio_final is None:
                    precio_final = matricula.costo_curso

                anticipo = matricula.primer_pago or Decimal("0.00")
                nuevo_saldo = precio_final - anticipo - total_pagado_cuotas
                if nuevo_saldo < 0:
                    nuevo_saldo = Decimal("0.00")

                if matricula.saldo_pendiente != nuevo_saldo:
                    matricula.saldo_pendiente = nuevo_saldo
                    matricula.save(update_fields=["saldo_pendiente"])

                # ✅ NUEVO: Reprogramaciones pendientes (cobros extras)
                reprogramaciones_pendientes = Reprogramacion.objects.filter(
                    matricula=matricula,
                    pagado=False,
                    monto__gt=0
                ).order_by("-creado_en")

                total_reprogramaciones_pendientes = (
                    reprogramaciones_pendientes.aggregate(t=Sum("monto"))["t"] or Decimal("0.00")
                )

                # ✅ NUEVO: saldo total mostrado (curso + reprogramaciones)
                saldo_con_reprogramaciones = (matricula.saldo_pendiente or Decimal("0.00")) + total_reprogramaciones_pendientes

    return render(request, "cursos/pagos.html", {
        "dni": dni,
        "alumno": alumno,
        "matriculas": matriculas,
        "matricula": matricula,
        "matricula_id": str(matricula.id) if matricula else "",
        "cuotas": cuotas,
        "total_pagado_cuotas": total_pagado_cuotas,

        # ✅ NUEVO
        "reprogramaciones_pendientes": reprogramaciones_pendientes,
        "total_reprogramaciones_pendientes": total_reprogramaciones_pendientes,
        "saldo_con_reprogramaciones": saldo_con_reprogramaciones,
    })

def crear_cuotas(matricula):
    saldo = matricula.precio_final - matricula.primer_pago

    if saldo <= 0:
        return

    monto_cuota = saldo / Decimal('2')

    # Cuota 1 (15 días después)
    Cuota.objects.create(
        matricula=matricula,
        numero=1,
        monto=monto_cuota,
        fecha_vencimiento=matricula.fecha_inicio + timedelta(days=15),
        pagado=False
    )

    # Cuota 2 (30 días después)
    Cuota.objects.create(
        matricula=matricula,
        numero=2,
        monto=monto_cuota,
        fecha_vencimiento=matricula.fecha_inicio + timedelta(days=30),
        pagado=False
    )

@login_required
def pagar_cuota(request, cuota_id):
    if request.method != "POST":
        return redirect("pagina_pagos")

    cuota = get_object_or_404(Cuota, id=cuota_id)

    dni = (request.POST.get("dni") or "").strip()
    matricula_id = (request.POST.get("matricula_id") or "").strip()

    monto_str = (request.POST.get("monto") or "").strip()
    if not monto_str:
        messages.error(request, "Ingresa un monto.")
        return redirect(f"/cursos/pagos/?dni={dni}&matricula_id={matricula_id}")

    try:
        monto = Decimal(monto_str)
    except:
        messages.error(request, "Monto inválido.")
        return redirect(f"/cursos/pagos/?dni={dni}&matricula_id={matricula_id}")

    if monto <= 0:
        messages.error(request, "El monto debe ser mayor a 0.")
        return redirect(f"/cursos/pagos/?dni={dni}&matricula_id={matricula_id}")

    saldo_cuota = (cuota.monto - (cuota.monto_pagado or Decimal("0.00")))
    if monto > saldo_cuota:
        monto = saldo_cuota

    # ✅ NUEVO: obtener método de pago desde el form
    metodo_pago = (request.POST.get("metodo_pago") or "").strip().lower()

    # ✅ Validar contra choices del modelo
    metodos_validos = dict(Pago.METODO_PAGO_CHOICES).keys()
    if metodo_pago not in metodos_validos:
        messages.error(request, "Selecciona un método de pago válido.")
        return redirect(f"/cursos/pagos/?dni={dni}&matricula_id={matricula_id}")

    # ✅ registrar el pago
    Pago.objects.create(
        matricula=cuota.matricula,
        cuota=cuota,
        monto=monto,
        concepto=f"cuota_{cuota.numero}",
        registrado_por=request.user,
        metodo_pago=metodo_pago,
        fecha_pago_real=timezone.now().date(),
    )

    # ✅ actualizar cuota
    cuota.monto_pagado = (cuota.monto_pagado or Decimal("0.00")) + monto
    if cuota.monto_pagado >= cuota.monto:
        cuota.pagado = True
    cuota.save()

    # ✅ recalcular saldo pendiente usando SOLO pagos activos
    matricula = cuota.matricula
    total_pagado = (
        Pago.objects
        .filter(matricula=matricula, activo=True)
        .aggregate(s=Sum("monto"))["s"]
        or Decimal("0.00")
    )

    # ✅ usa precio_final si existe
    precio_final = getattr(matricula, "precio_final", None) or matricula.costo_curso
    matricula.saldo_pendiente = precio_final - total_pagado
    if matricula.saldo_pendiente < 0:
        matricula.saldo_pendiente = Decimal("0.00")
    matricula.save(update_fields=["saldo_pendiente"])

    ruta_cert = verificar_y_generar_certificado(matricula)
    if ruta_cert:
        messages.success(request, "✅ Pago completo y asistencias completas: Certificado generado.")
    else:
        messages.success(request, "✅ Pago registrado correctamente.")

    return redirect(f"/cursos/pagos/?dni={dni}&matricula_id={cuota.matricula.id}")

@login_required
@solo_asesora
def egresados(request):
    todas = Matricula.objects.select_related("alumno", "curso")
    matriculas = [m for m in todas if puede_generar_certificado(m)]

    certificados = Certificado.objects.filter(matricula__in=matriculas)
    cert_map = {c.matricula_id: c for c in certificados}

    for m in matriculas:
        m.certificado = cert_map.get(m.id)

    return render(request, "cursos/egresados.html", {"matriculas": matriculas})

def asegurar_cuotas(matricula: Matricula):
    """
    Crea Cuota 1 y Cuota 2 si la matrícula no tiene cuotas.
    """
    existe = Cuota.objects.filter(matricula=matricula).exists()
    if existe:
        return

    # Precio final (si no existe, usa costo_curso)
    precio_final = getattr(matricula, "precio_final", None)
    if precio_final is None:
        precio_final = matricula.costo_curso

    anticipo = matricula.primer_pago or Decimal("0.00")
    saldo = precio_final - anticipo
    if saldo <= 0:
        return

    monto_cuota = (saldo / Decimal("2")).quantize(Decimal("0.01"))

    fecha_base = matricula.fecha_inicio or matricula.fecha_inscripcion
    Cuota.objects.create(
        matricula=matricula,
        numero=1,
        monto=monto_cuota,
        monto_pagado=Decimal("0.00"),
        pagado=False,
        fecha_vencimiento=fecha_base + timedelta(days=15),
    )
    Cuota.objects.create(
        matricula=matricula,
        numero=2,
        monto=monto_cuota,
        monto_pagado=Decimal("0.00"),
        pagado=False,
        fecha_vencimiento=fecha_base + timedelta(days=30),
    )
def wrap_center(text, font_name, font_size, max_width):
    words = text.split()
    lines, line = [], ""
    for w in words:
        test = (line + " " + w).strip()
        if stringWidth(test, font_name, font_size) <= max_width:
            line = test
        else:
            if line:
                lines.append(line)
            line = w
    if line:
        lines.append(line)
    return lines

@login_required
def generar_certificado(request, matricula_id):
    matricula = get_object_or_404(Matricula, id=matricula_id)
    alumno = matricula.alumno

    if not alumno.correo:
        messages.error(request, "❌ El alumno no tiene correo registrado.")
        return redirect("egresados")

    force_regen = request.GET.get("regen") == "1"

    cert = verificar_y_generar_certificado(matricula, force_regen=force_regen)
    if not cert or not cert.archivo_pdf:
        messages.error(request, "❌ No se pudo generar el certificado (no cumple requisitos o faltan datos).")
        return redirect("egresados")

    pdf_path = cert.archivo_pdf.path  # ✅ ahora sí es un path real

    asunto = "🎓 Tu certificado - WeeinClass"
    mensaje = f"""
Hola {alumno.nombre},

Felicitaciones por culminar tu curso:

📚 {matricula.curso.nombre}

Adjuntamos tu certificado oficial.

Gracias por confiar en WeeinClass 💙

Saludos,
Equipo WeeinClass
    """

    email = EmailMessage(
        subject=asunto,
        body=mensaje,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[alumno.correo],
    )

    email.attach_file(pdf_path)

    try:
        email.send(fail_silently=False)
        messages.success(request, f"✅ Certificado enviado a {alumno.correo}")
    except Exception as e:
        messages.error(request, f"❌ Error enviando correo: {str(e)}")

    return redirect("egresados")

@login_required
def pagar_reprogramacion(request):
    reprog_id = (request.GET.get("reprogramacion_id") or "").strip()
    if not reprog_id:
        messages.error(request, "No se encontró la reprogramación.")
        return redirect("pagina_pagos")

    reprog = get_object_or_404(Reprogramacion, id=reprog_id)
    matricula = reprog.matricula

    # ✅ Si ya está pagada, o no requiere cobro, no tiene sentido confirmar
    if reprog.pagado or (reprog.monto is not None and reprog.monto <= 0):
        messages.info(request, "Esta reprogramación no tiene cobro pendiente.")
        return redirect(f"/cursos/pagos/?dni={matricula.alumno.dni}&matricula_id={matricula.id}")

    # ✅ SOLO CONFIRMACIÓN: aquí NO se cobra
    if request.method == "POST":
        # asegurar monto correcto
        if not reprog.monto or reprog.monto <= 0:
            reprog.monto = Decimal("50.00")
        else:
            reprog.monto = Decimal(reprog.monto).quantize(Decimal("0.01"))

        # dejar pendiente (cobro por gestionar en /cursos/pagos/)
        reprog.pagado = False
        reprog.pagado_en = None

        # asesora que confirmó el cobro pendiente
        reprog.creado_por = request.user

        reprog.save(update_fields=["monto", "pagado", "pagado_en", "creado_por"])

        messages.warning(
            request,
            "⚠ Cobro pendiente por reprogramación registrado. Ve a Gestión de Pagos para pagarlo."
        )

        return redirect(f"/cursos/pagos/?dni={matricula.alumno.dni}&matricula_id={matricula.id}")

    return render(request, "cursos/confirmar_reprogramacion.html", {
        "matricula": matricula,
        "reprogramacion": reprog,
    })

@require_POST
@login_required
def pagar_reprogramacion_pagos(request, reprog_id):
    reprog = get_object_or_404(Reprogramacion, id=reprog_id)
    matricula = reprog.matricula
    clase = reprog.clase

    dni = (request.POST.get("dni") or matricula.alumno.dni).strip()
    matricula_id = (request.POST.get("matricula_id") or str(matricula.id)).strip()

    # ✅ validar método
    metodo_pago = (request.POST.get("metodo_pago") or "").strip().lower()
    metodos_validos = dict(Pago.METODO_PAGO_CHOICES).keys()
    if metodo_pago not in metodos_validos:
        messages.error(request, "Selecciona un método de pago válido.")
        return redirect(f"/cursos/pagos/?dni={dni}&matricula_id={matricula_id}")

    if reprog.pagado:
        messages.info(request, "Esta reprogramación ya fue pagada.")
        return redirect(f"/cursos/pagos/?dni={dni}&matricula_id={matricula_id}")

    if reprog.monto is None or reprog.monto <= 0:
        reprog.monto = Decimal("50.00")

    try:
        with transaction.atomic():
            # ✅ Registrar pago en historial
            Pago.objects.create(
                matricula=matricula,
                cuota=None,
                monto=reprog.monto,
                fecha_pago_real=timezone.now().date(),
                metodo_pago=metodo_pago,
                concepto="reprogramacion",
                registrado_por=request.user,
                activo=True
            )

            # ✅ Aplicar el cambio recién cuando se pague
            clase.fecha = reprog.fecha_nueva
            clase.save(update_fields=["fecha"])

            # ✅ Marcar reprogramación pagada
            reprog.pagado = True
            reprog.pagado_en = timezone.now()
            reprog.creado_por = request.user
            reprog.save(update_fields=["pagado", "pagado_en", "creado_por"])

        messages.success(request, "✅ Reprogramación pagada y aplicada.")
    except Exception as e:
        messages.error(request, f"❌ Error al pagar reprogramación: {e}")

    return redirect(f"/cursos/pagos/?dni={dni}&matricula_id={matricula_id}")

@transaction.atomic
def asegurar_unidades_curso(curso):
    codigo = (getattr(curso, "codigo", "") or "").strip().upper()

    if not codigo:
        nombre = (getattr(curso, "nombre", "") or "").strip()
        codigo, _ = curso_codigo_y_sesiones(nombre, curso_obj=curso)

    total = CURSO_SESIONES.get(codigo, 6)

    # Temas del curso (si no hay, cae a "Sesión X")
    temas = TEMAS_POR_CODIGO.get(codigo, [])

    # Limpiar duplicados por numero
    unidades = UnidadCurso.objects.filter(curso=curso).order_by("numero", "id")
    vistos = set()
    for u in unidades:
        if u.numero in vistos:
            u.delete()
        else:
            vistos.add(u.numero)

    # Crear/actualizar unidades con nombre correcto
    for n in range(1, total + 1):
        tema = temas[n - 1] if (n - 1) < len(temas) else None

        if tema:
            nombre_tema = f"Sesión {n} - {tema}"
        else:
            nombre_tema = f"Sesión {n}"

        obj, created = UnidadCurso.objects.get_or_create(
            curso=curso,
            numero=n,
            defaults={"nombre_tema": nombre_tema}
        )

        # ✅ si ya existía, actualizarlo también
        if not created and obj.nombre_tema != nombre_tema:
            obj.nombre_tema = nombre_tema
            obj.save(update_fields=["nombre_tema"])

    # borrar sobrantes
    UnidadCurso.objects.filter(curso=curso, numero__gt=total).delete()

    return list(UnidadCurso.objects.filter(curso=curso).order_by("numero"))

def get_codigo_curso(curso):
    # toma el código de la primera unidad del curso
    u = UnidadCurso.objects.filter(curso=curso).order_by("numero").first()
    return (u.codigo or "").strip().upper() if u else ""

from decimal import Decimal
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from cursos.models import Alumno, Matricula, Pago


@login_required
def historial_pagos(request):
    dni = (request.GET.get("dni") or "").strip()
    matricula_id = (request.GET.get("matricula_id") or "").strip()
    export = (request.GET.get("export") or "").strip()

    alumno = None
    matriculas = Matricula.objects.none()
    matricula = None

    pagos_matricula = Pago.objects.none()
    pagos_servicios = Pago.objects.none()

    total_general = Decimal("0.00")
    total_curso = Decimal("0.00")
    total_servicios = Decimal("0.00")

    if dni:
        alumno = Alumno.objects.filter(dni=dni).first()

        if alumno:

            matriculas = Matricula.objects.filter(
                alumno=alumno
            ).order_by("-fecha_inscripcion")

            # Matrícula seleccionada
            if matricula_id:
                matricula = matriculas.filter(id=matricula_id).first()

            if not matricula and matriculas.exists():
                matricula = matriculas.first()

            # ===============================
            # PAGOS DE MATRÍCULA
            # ===============================
            if matricula:

                pagos_matricula = Pago.objects.filter(
                    activo=True,
                    matricula=matricula
                ).select_related(
                    "registrado_por",
                    "matricula__curso"
                ).order_by("-fecha_pago_real", "-creado_en")

                total_curso = pagos_matricula.aggregate(
                    t=Sum("monto")
                )["t"] or Decimal("0.00")

            # ===============================
            # PAGOS DE SERVICIOS
            # ===============================
            pagos_servicios = Pago.objects.filter(
                activo=True,
                alumno=alumno,
                matricula__isnull=True,
                concepto="servicio_extra"
            ).select_related(
                "registrado_por"
            ).order_by("-fecha_pago_real", "-creado_en")

            total_servicios = pagos_servicios.aggregate(
                t=Sum("monto")
            )["t"] or Decimal("0.00")

            # ===============================
            # TOTAL GENERAL
            # ===============================
            total_general = total_curso + total_servicios

    return render(request, "cursos/historial_pagos.html", {
        "dni": dni,
        "alumno": alumno,
        "matriculas": matriculas,
        "matricula": matricula,

        "pagos_matricula": pagos_matricula,
        "pagos_servicios": pagos_servicios,

        "total_curso": total_curso,
        "total_servicios": total_servicios,
        "total_general": total_general,
    })


@login_required
def exportar_historial_pagos_excel(request):
    dni = (request.GET.get("dni") or "").strip()
    matricula_id = (request.GET.get("matricula_id") or "").strip()

    if not dni:
        # export vacío (o puedes messages + redirect)
        return exportar_excel(
            "Historial de pagos (sin DNI)",
            ["Fecha", "Alumno", "Curso", "Concepto", "Método", "Monto", "Asesora"],
            [],
            filename="historial_pagos.xlsx"
        )

    alumno = Alumno.objects.filter(dni=dni).first()
    if not alumno:
        return exportar_excel(
            f"Historial de pagos - DNI {dni} (no encontrado)",
            ["Fecha", "Alumno", "Curso", "Concepto", "Método", "Monto", "Asesora"],
            [],
            filename=f"historial_pagos_{dni}.xlsx"
        )

    matriculas = Matricula.objects.filter(alumno=alumno).order_by("-fecha_inscripcion", "-id")
    matricula = None
    if matricula_id:
        matricula = matriculas.filter(id=matricula_id).first()
    if not matricula:
        matricula = matriculas.first()

    pagos = (
        Pago.objects
        .filter(matricula=matricula, activo=True)
        .select_related("matricula__curso", "matricula__alumno", "registrado_por")
        .order_by("-fecha_pago_real", "-id")
    )

    filas = []
    for p in pagos:
        filas.append([
            str(p.fecha_pago_real or ""),
            p.matricula.alumno.nombre,
            p.matricula.curso.nombre,
            p.get_concepto_display() if hasattr(p, "get_concepto_display") else p.concepto,
            p.get_metodo_pago_display() if p.metodo_pago else "",
            float(p.monto),
            (p.registrado_por.username if p.registrado_por else ""),
        ])

    columnas = ["Fecha", "Alumno", "Curso", "Concepto", "Método", "Monto", "Asesora"]

    titulo = f"Historial de pagos - {alumno.nombre} (DNI: {dni}) - {matricula.curso.nombre}"

    return exportar_excel(
        titulo,
        columnas,
        filas,
        filename=f"historial_pagos_{dni}_matricula_{matricula.id}.xlsx"
    )

@login_required
def catalogo_servicios(request):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para vender servicios.")
        return redirect("menu_admin")

    dni = (request.GET.get("dni") or request.POST.get("dni") or "").strip()

    alumno = None
    servicios = ServicioExtra.objects.filter(activo=True).order_by("nombre")
    total_servicios = Decimal("0.00")
    compras = Pago.objects.none()

    # ✅ detectar si Pago tiene el campo alumno (por seguridad)
    try:
        Pago._meta.get_field("alumno")
        tiene_alumno = True
    except Exception:
        tiene_alumno = False

    if dni:
        alumno = Alumno.objects.filter(dni=dni).first()
        if alumno and tiene_alumno:
            compras = (
                Pago.objects
                .filter(alumno=alumno, concepto="servicio_extra", activo=True, matricula__isnull=True)
                .select_related("registrado_por")
                .order_by("-fecha_pago_real", "-creado_en", "-id")
            )
            total_servicios = compras.aggregate(s=Sum("monto"))["s"] or Decimal("0.00")

    if request.method == "POST":
        if not dni:
            messages.error(request, "Ingresa un DNI.")
            return redirect("catalogo_servicios")

        alumno = Alumno.objects.filter(dni=dni).first()
        if not alumno:
            messages.error(request, "No se encontró alumno con ese DNI.")
            return redirect(f"{request.path}?dni={dni}")

        if not tiene_alumno:
            messages.error(request, "Tu modelo Pago no tiene el campo alumno. Falta aplicar el cambio del modelo.")
            return redirect(f"{request.path}?dni={dni}")

        servicio_id = (request.POST.get("servicio_id") or "").strip()
        metodo_pago = (request.POST.get("metodo_pago") or "").strip().lower()
        cantidad_str = (request.POST.get("cantidad") or "1").strip()

        if not servicio_id:
            messages.error(request, "Selecciona un servicio.")
            return redirect(f"{request.path}?dni={dni}")

        try:
            cantidad = int(cantidad_str)
            if cantidad <= 0:
                raise ValueError()
        except Exception:
            messages.error(request, "Cantidad inválida.")
            return redirect(f"{request.path}?dni={dni}")

        metodos_validos = dict(Pago.METODO_PAGO_CHOICES).keys()
        if metodo_pago not in metodos_validos:
            messages.error(request, "Selecciona un método de pago válido.")
            return redirect(f"{request.path}?dni={dni}")

        servicio = get_object_or_404(ServicioExtra, id=servicio_id, activo=True)
        total = (servicio.precio * Decimal(cantidad)).quantize(Decimal("0.01"))

        # ✅ Registrar compra como Pago "sin matrícula" pero ligado a alumno
        Pago.objects.create(
            matricula=None,
            alumno=alumno,
            monto=total,
            fecha_pago_real=timezone.now().date(),
            metodo_pago=metodo_pago,
            concepto="servicio_extra",
            detalle=f"{servicio.nombre} x{cantidad}",
            registrado_por=request.user,
            activo=True
        )

        messages.success(request, f"✅ Cobrado: {servicio.nombre} x{cantidad} (S/ {total}).")
        return redirect(f"{request.path}?dni={dni}")

    return render(request, "cursos/catalogo_servicios.html", {
        "dni": dni,
        "alumno": alumno,
        "servicios": servicios,
        "compras": compras,
        "total_servicios": total_servicios,
    })


def editar_servicio(request, servicio_id):
    servicio = get_object_or_404(ServicioExtra, id=servicio_id)
    form = ServicioForm(request.POST or None, instance=servicio)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "✅ Servicio actualizado.")
        return redirect("catalogo_servicios")

    return render(request, "cursos/servicio_editar.html", {"form": form, "servicio": servicio})

def _es_admin(user):
    return (
        user.is_superuser
        or user.groups.filter(name__in=["Admin", "Administradores"]).exists()
    )

@login_required
def servicios_listar_crear(request):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para gestionar servicios.")
        return redirect("menu_admin")

    q = (request.GET.get("q") or "").strip()

    servicios = ServicioExtra.objects.all().order_by("-activo", "nombre")
    if q:
        servicios = servicios.filter(Q(nombre__icontains=q))

    # ✅ Crear
    if request.method == "POST":
        form = ServicioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Servicio creado correctamente.")
            return redirect("servicios_gestion")
        else:
            messages.error(request, "Revisa los campos del servicio.")
    else:
        form = ServicioForm()

    return render(request, "cursos/servicios_gestion.html", {
        "form": form,
        "servicios": servicios,
        "q": q,
    })


@login_required
def servicios_editar(request, servicio_id):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para editar servicios.")
        return redirect("menu_admin")

    servicio = get_object_or_404(ServicioExtra, id=servicio_id)

    if request.method == "POST":
        form = ServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Servicio actualizado.")
            return redirect("servicios_gestion")
        else:
            messages.error(request, "Revisa los campos.")
    else:
        form = ServicioForm(instance=servicio)

    return render(request, "cursos/servicio_editar.html", {
        "form": form,
        "servicio": servicio,
    })


@login_required
def servicios_toggle_activo(request, servicio_id):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para cambiar el estado.")
        return redirect("menu_admin")

    if request.method != "POST":
        return redirect("servicios_gestion")

    servicio = get_object_or_404(ServicioExtra, id=servicio_id)
    servicio.activo = not servicio.activo
    servicio.save(update_fields=["activo"])

    if servicio.activo:
        messages.success(request, "✅ Servicio activado.")
    else:
        messages.warning(request, "⚠ Servicio desactivado (no se podrá vender).")

    return redirect("servicios_gestion")

@require_POST
@login_required
def toggle_kit_matricula(request, matricula_id):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos.")
        return redirect("lista_matriculas")

    tipo = (request.POST.get("tipo") or "").strip()  # "manual" o "video"
    matricula = get_object_or_404(Matricula, id=matricula_id)

    kit, _ = EntregaKit.objects.get_or_create(matricula=matricula)

    ahora = timezone.now()

    if tipo == "manual":
        kit.manual_entregado = not kit.manual_entregado
        kit.manual_entregado_en = ahora if kit.manual_entregado else None
    elif tipo == "video":
        kit.video_enviado = not kit.video_enviado
        kit.video_enviado_en = ahora if kit.video_enviado else None
    else:
        messages.error(request, "Acción inválida.")
        return redirect("lista_matriculas")

    kit.actualizado_por = request.user
    kit.save()

    messages.success(request, "✅ Estado del kit actualizado.")
    return redirect("lista_matriculas")

@login_required
def reporte_caja(request):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para ver el reporte de caja.")
        return redirect("menu_admin")

    # filtros
    fecha_ini = (request.GET.get("fecha_ini") or "").strip()  # YYYY-MM-DD
    fecha_fin = (request.GET.get("fecha_fin") or "").strip()  # YYYY-MM-DD
    metodo = (request.GET.get("metodo") or "").strip().lower()
    export = (request.GET.get("export") or "").strip()  # "1" excel

    pagos = Pago.objects.filter(activo=True)

    # rango por fecha real (si no existe, usa creado_en)
    if fecha_ini:
        pagos = pagos.filter(Q(fecha_pago_real__gte=fecha_ini) | Q(fecha_pago_real__isnull=True, creado_en__date__gte=fecha_ini))
    if fecha_fin:
        pagos = pagos.filter(Q(fecha_pago_real__lte=fecha_fin) | Q(fecha_pago_real__isnull=True, creado_en__date__lte=fecha_fin))

    if metodo:
        pagos = pagos.filter(metodo_pago=metodo)

    pagos = pagos.select_related(
        "registrado_por",
        "matricula",
        "matricula__alumno",
        "matricula__curso",
        "cuota"
    ).order_by("-fecha_pago_real", "-creado_en", "-id")

    # totales
    total = pagos.aggregate(t=Sum("monto"))["t"] or Decimal("0.00")

    por_metodo = (
        pagos.values("metodo_pago")
        .annotate(total=Sum("monto"))
        .order_by("-total")
    )

    por_concepto = (
        pagos.values("concepto")
        .annotate(total=Sum("monto"))
        .order_by("-total")
    )

    # export excel
    if export == "1":
        wb = Workbook()
        ws = wb.active
        ws.title = "Caja"

        headers = ["Fecha pago", "Registrado en", "Alumno (DNI)", "Curso", "Concepto", "Método", "Monto", "Asesora"]
        ws.append(headers)

        for p in pagos:
            # alumno
            alumno_nombre = "—"
            alumno_dni = "—"
            curso = "Servicio (sin curso)"

            if p.matricula_id:
                if p.matricula.alumno:
                    alumno_nombre = p.matricula.alumno.nombre
                    alumno_dni = p.matricula.alumno.dni
                if p.matricula.curso:
                    curso = p.matricula.curso.nombre

            # método
            metodo_txt = p.get_metodo_pago_display() if p.metodo_pago else "—"

            # asesora
            asesora = "—"
            if p.registrado_por:
                asesora = (p.registrado_por.get_full_name() or p.registrado_por.username)

            ws.append([
                p.fecha_pago_real.strftime("%d/%m/%Y") if p.fecha_pago_real else "—",
                p.creado_en.strftime("%d/%m/%Y %H:%M") if p.creado_en else "—",
                f"{alumno_nombre} ({alumno_dni})" if alumno_nombre != "—" else "—",
                curso,
                p.get_concepto_display() if p.concepto else "—",
                metodo_txt,
                float(p.monto),
                asesora,
            ])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fname = f"reporte_caja_{timezone.now().date()}.xlsx"
        resp = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    return render(request, "cursos/reporte_caja.html", {
        "fecha_ini": fecha_ini,
        "fecha_fin": fecha_fin,
        "metodo": metodo,
        "pagos": pagos,
        "total": total,
        "por_metodo": por_metodo,
        "por_concepto": por_concepto,
    })

def rango_periodo(periodo: str):
    hoy = timezone.localdate()

    if periodo == "semana":
        # lunes a domingo (semana actual)
        fi = hoy - timezone.timedelta(days=hoy.weekday())
        ff = fi + timezone.timedelta(days=6)
        return fi, ff

    if periodo == "anio":
        fi = hoy.replace(month=1, day=1)
        ff = hoy
        return fi, ff

    # default: mes
    fi = hoy.replace(day=1)
    ff = hoy
    return fi, ff

def alumnos_por_curso_periodo(curso_id=None, periodo="mes"):
    cursos_qs = Curso.objects.all()
    if curso_id:
        cursos_qs = cursos_qs.filter(id=curso_id)

    fi, ff = rango_periodo(periodo)

    matriculas = Matricula.objects.filter(curso__in=cursos_qs)
    if fi and ff:
        matriculas = matriculas.filter(fecha_inscripcion__range=[fi, ff])

    data = (
        cursos_qs
        .annotate(
            total_alumnos=Count(
                "matricula",
                filter=Q(matricula__in=matriculas),
                distinct=True
            )
        )
        .values("nombre", "total_alumnos")
    )
    return list(data)

@login_required
def marcar_truncado(request, matricula_id):
    m = get_object_or_404(Matricula.objects.select_related("alumno", "curso"), id=matricula_id)

    if request.method == "POST":
        motivo = (request.POST.get("motivo_trunco") or "").strip()
        razon = (request.POST.get("razon_trunco") or "").strip()
        fecha_baja = request.POST.get("fecha_baja") or None

        m.estado = "truncado"
        m.motivo_trunco = motivo or None
        m.razon_trunco = razon
        m.fecha_baja = fecha_baja or timezone.now().date()
        m.save()

        messages.success(request, "✅ Matrícula marcada como TRUNCADO.")
        return redirect("lista_matriculas")

    return render(request, "cursos/marcar_truncado.html", {"m": m})

@login_required
def alumnos_no_terminaron(request):
    hoy = timezone.now().date()

    qs = Matricula.objects.select_related("alumno", "curso").exclude(estado="finalizada")

    # opcional: solo los que ya iniciaron
    qs = qs.filter(fecha_inicio__isnull=False, fecha_inicio__lte=hoy)

    return render(request, "cursos/no_terminaron.html", {"matriculas": qs})

@login_required
def crear_usuario(request):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para crear usuarios.")
        return redirect("menu_admin")

    if request.method == "POST":
        form = UsuarioCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Usuario creado correctamente.")
            return redirect("lista_usuarios")
    else:
        form = UsuarioCreateForm()

    return render(request, "cursos/crear_usuario.html", {"form": form})

@login_required
def lista_usuarios(request):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para ver usuarios.")
        return redirect("menu_admin")

    q = (request.GET.get("q") or "").strip()

    usuarios = User.objects.all().order_by("username")
    if q:
        usuarios = usuarios.filter(
            username__icontains=q
        ) | usuarios.filter(
            first_name__icontains=q
        ) | usuarios.filter(
            last_name__icontains=q
        ) | usuarios.filter(
            email__icontains=q
        )

    return render(request, "cursos/lista_usuarios.html", {
        "usuarios": usuarios,
        "q": q
    })

@login_required
def editar_usuario(request, user_id):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para editar usuarios.")
        return redirect("menu_admin")

    usuario = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        form = UsuarioUpdateForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Usuario actualizado.")
            return redirect("lista_usuarios")
    else:
        form = UsuarioUpdateForm(instance=usuario)

    return render(request, "cursos/editar_usuario.html", {
        "form": form,
        "usuario": usuario
    })

@login_required
def eliminar_usuario(request, user_id):
    if not _es_admin(request.user):
        messages.error(request, "No tienes permisos para eliminar usuarios.")
        return redirect("menu_admin")

    usuario = get_object_or_404(User, id=user_id)

    # evita borrar tu propio usuario por accidente
    if usuario.id == request.user.id:
        messages.error(request, "No puedes eliminar tu propio usuario.")
        return redirect("lista_usuarios")

    if request.method == "POST":
        usuario.delete()
        messages.success(request, "🗑 Usuario eliminado.")
        return redirect("lista_usuarios")

    return render(request, "cursos/eliminar_usuario.html", {"usuario": usuario})